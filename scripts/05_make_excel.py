"""
05_make_excel.py — Build 9-sheet Excel workbook in v4 reference style.

Input:  ~/german-poetry-v5/output/poems_v5.csv
Output: ~/german-poetry-v5/output/german_poetry_v5.xlsx

Sheets: Poems | By Author | By Collection | Statistical Tests |
        Correlation | Temporal Confound | Top Outliers | Notes | Normalization
"""

from pathlib import Path
import pandas as pd
import numpy as np
from scipy import stats as sp_stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

OUT_DIR = Path.home() / "german-poetry-v5" / "output"

# ── Load data ─────────────────────────────────────────────────────────────────
df = pd.read_csv(OUT_DIR / "poems_v5.csv")

# Decade bins (temporal_year = erstdruck when available, else comp_year if not uncertain)
df["temporal_decade"] = (df["temporal_year"] // 10 * 10).where(df["temporal_year"].notna())
df["mean_stanza_len"] = (df["n_lines"] / df["n_stanzas"].replace(0, np.nan)).round(1)

METRICS = ["mean_surprisal", "uid_sigma", "s2_mean", "mean_entropy",
           "tension", "peak_pos", "ac1", "ttr"]
METRIC_LABELS = {
    "mean_surprisal": "Mean Surprisal",
    "uid_sigma":      "UID σ",
    "s2_mean":        "S² Mean",
    "mean_entropy":   "Mean Entropy",
    "tension":        "Tension",
    "peak_pos":       "Peak Pos",
    "ac1":            "AC1",
    "ttr":            "TTR",
}

# ── Colour palette (matching v4 reference) ────────────────────────────────────
def fill(hex6): return PatternFill("solid", fgColor=hex6)

C_META   = fill("1A5276")   # metadata (dark blue)
C_DATE   = fill("117A65")   # dates (green)
C_STRUCT = fill("1A5276")   # structural counts
C_ENTR   = fill("154360")   # entropic metrics (navy)
C_ANN    = fill("4A235A")   # annotations (purple)
C_DERIV  = fill("424949")   # derived/extra (grey)
C_BANNER = fill("0D0D0D")   # banner row

HEADER_FONT  = Font(color="E6B85A", bold=True, size=10)
BANNER_FONT  = Font(color="E6B85A", bold=True, size=11, italic=True)
DATA_FONT    = Font(size=9)

POET_COLORS = {
    "trakl":                    "C8E6C9",
    "heym":                     "FFCDD2",
    "stadler":                  "FFE0B2",
    "hvh":                      "F8BBD9",
    "Felix Dörmann":            "E1BEE7",
    "Morgenstern":              "B3E5FC",
    "Rilke":                    "DCEDC8",
    "Holderlin":                "FFF9C4",
    "Mörike":                   "F0F4C3",
    "storm":                    "B2DFDB",
    "goethe":                   "CFD8DC",
    "Schiller":                 "D7CCC8",
    "Klopstock":                "BBDEFB",
    "Droste-Hülshoff":          "F3E5F5",
    "Barthold Heinrich Brockes":"E8F5E9",
    "Albrech von Haller":       "FBE9E7",
    "wedekind":                 "EDE7F6",
    "george":                   "FFF3E0",
}

def set_header(cell, col_fill, val):
    cell.value = val
    cell.fill  = col_fill
    cell.font  = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

def autowidth(ws, max_w=45, min_w=6):
    for col in ws.columns:
        try:
            length = max(len(str(c.value or "")) for c in col if c.row <= 200)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(min(length + 2, max_w), min_w)
        except Exception:
            pass

def section_header(ws, row, col, text, bgcolor="2C3E50", fgcolor="ECF0F1"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=bgcolor)
    cell.font = Font(color=fgcolor, bold=True, size=10)
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Poems
# ═══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "Poems"
ws.freeze_panes = "A3"

POEMS_COLS = [
    # (header_label, data_column, fill_color)
    ("Poet",              "poet",               C_META),
    ("Movement",          "movement",           C_META),
    ("Collection",        "collection",         C_META),
    ("Poem Title",        "poem_title",         C_META),
    ("Comp Year",         "comp_year",          C_DATE),
    ("Erstdruck Year",    "erstdruck_year_int", C_DATE),
    ("Edition Year",      "edition_year_int",   C_DATE),
    ("Temporal Year",     "temporal_year",      C_DATE),
    ("Date Uncertain",    "date_uncertain",     C_DATE),
    ("Lines",             "n_lines",            C_STRUCT),
    ("Stanzas",           "n_stanzas",          C_STRUCT),
    ("Mean Stanza Len",   "mean_stanza_len",    C_STRUCT),
    ("Token Count",       "n_words",            C_STRUCT),
    ("Mean Surprisal",    "mean_surprisal",     C_ENTR),
    ("UID σ",             "uid_sigma",          C_ENTR),
    ("S² Mean",           "s2_mean",            C_ENTR),
    ("Mean Entropy",      "mean_entropy",       C_ENTR),
    ("Tension",           "tension",            C_ENTR),
    ("Peak Pos",          "peak_pos",           C_ENTR),
    ("AC1",               "ac1",                C_ENTR),
    ("TTR",               "ttr",                C_ENTR),
    ("Person",            "person",             C_ANN),
    ("Addressee",         "has_addressee",      C_ANN),
    ("Imagery",           "imagery_density",    C_ANN),
    ("Valence",           "emotional_valence",  C_ANN),
    ("Intensity",         "emotional_intensity",C_ANN),
    ("Tone",              "tone",               C_ANN),
    ("Time Frame",        "temporal_frame",     C_ANN),
    ("Setting",           "setting",            C_ANN),
    ("Closure",           "closure",            C_ANN),
    ("Text Variant",      "_text_variant",      C_DERIV),
    ("Text",              "_text_display",      C_DERIV),
]

# Build row list: non-George = 1 row each; George = 2 rows (original then normalized)
poem_rows_list = []
for _, row_data in df.iterrows():
    r = row_data.to_dict()
    poet = str(r.get("poet", ""))
    if poet == "george":
        orig = dict(r); orig["_text_variant"] = "original";   orig["_text_display"] = r.get("text", "")
        norm = dict(r); norm["_text_variant"] = "normalized"; norm["_text_display"] = r.get("george_normalized_text", "")
        # Metrics only on original row; blank them on normalized
        for m in METRICS:
            norm[m] = ""
        poem_rows_list.append(orig)
        poem_rows_list.append(norm)
    else:
        r["_text_variant"] = ""
        r["_text_display"]  = ""
        poem_rows_list.append(r)

# Banner
ws.cell(row=1, column=1,
        value="Model: dbmdz/german-gpt2 | Annotations: claude-haiku-4-5-20251001 | Corpus: TextGrid / german-poetry-v5 | 2026-04-28"
        ).font = BANNER_FONT
ws.cell(row=1, column=1).fill = C_BANNER
ws.row_dimensions[1].height = 18

# Headers
for ci, (label, _, cf) in enumerate(POEMS_COLS, 1):
    set_header(ws.cell(row=2, column=ci), cf, label)
ws.row_dimensions[2].height = 30

# Data rows
for ri, row_data in enumerate(poem_rows_list, 3):
    poet = str(row_data.get("poet", ""))
    variant = row_data.get("_text_variant", "")
    fhex = POET_COLORS.get(poet, "FFFFFF")
    if poet == "george" and variant == "normalized":
        row_fill = fill("FFE0B2")   # slightly darker orange for normalized George rows
    else:
        row_fill = fill(fhex) if ri % 2 == 1 else fill("F8F8F8")

    for ci, (_, col_key, _) in enumerate(POEMS_COLS, 1):
        val = row_data.get(col_key, "")
        if pd.isna(val):
            val = ""
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.fill  = row_fill
        cell.font  = DATA_FONT
        cell.alignment = Alignment(wrap_text=False)

autowidth(ws)
ws.column_dimensions[get_column_letter(len(POEMS_COLS))].width = 70   # Text column


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: By Author
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("By Author")

agg_cols = {m: ["mean", "std"] for m in METRICS}
agg_cols["n_words"] = ["mean"]
poet_agg = df.groupby("poet").agg(
    n_poems=("poem_id", "count"),
    **{f"{m}_mean": (m, "mean") for m in METRICS},
    **{f"{m}_sd":   (m, "std")  for m in METRICS},
    token_mean=("n_words", "mean"),
).round(4).reset_index()
poet_agg = poet_agg.sort_values("mean_surprisal_mean", ascending=False)

by_author_cols = (
    [("Author", "poet", C_META), ("N", "n_poems", C_META)]
    + [(f"{METRIC_LABELS[m]} Mean", f"{m}_mean", C_ENTR) for m in METRICS]
    + [("Token Ct. Mean", "token_mean", C_STRUCT)]
    + [(f"{METRIC_LABELS[m]} SD", f"{m}_sd", C_ENTR) for m in METRICS]
)

for ci, (label, _, cf) in enumerate(by_author_cols, 1):
    set_header(ws2.cell(row=1, column=ci), cf, label)

for ri, (_, row_data) in enumerate(poet_agg.iterrows(), 2):
    poet = str(row_data.get("poet", ""))
    row_fill = fill(POET_COLORS.get(poet, "FFFFFF"))
    for ci, (_, col_key, _) in enumerate(by_author_cols, 1):
        val = row_data.get(col_key, "")
        if pd.isna(val): val = ""
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.fill = row_fill
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="right" if ci > 2 else "left")

autowidth(ws2)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: By Collection
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("By Collection")

coll_agg = df.groupby(["poet", "collection"]).agg(
    n_poems=("poem_id", "count"),
    **{f"{m}_mean": (m, "mean") for m in METRICS},
    **{f"{m}_sd":   (m, "std")  for m in METRICS},
    token_mean=("n_words", "mean"),
).round(4).reset_index()
coll_agg = coll_agg.sort_values(["poet", "mean_surprisal_mean"], ascending=[True, False])

by_coll_cols = (
    [("Author", "poet", C_META), ("Collection", "collection", C_META), ("N", "n_poems", C_META)]
    + [(f"{METRIC_LABELS[m]} Mean", f"{m}_mean", C_ENTR) for m in METRICS]
    + [("Token Ct. Mean", "token_mean", C_STRUCT)]
    + [(f"{METRIC_LABELS[m]} SD", f"{m}_sd", C_ENTR) for m in METRICS]
)

for ci, (label, _, cf) in enumerate(by_coll_cols, 1):
    set_header(ws3.cell(row=1, column=ci), cf, label)

for ri, (_, row_data) in enumerate(coll_agg.iterrows(), 2):
    poet = str(row_data.get("poet", ""))
    row_fill = fill(POET_COLORS.get(poet, "FFFFFF"))
    for ci, (_, col_key, _) in enumerate(by_coll_cols, 1):
        val = row_data.get(col_key, "")
        if pd.isna(val): val = ""
        cell = ws3.cell(row=ri, column=ci, value=val)
        cell.fill = row_fill
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="right" if ci > 3 else "left")

autowidth(ws3)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Statistical Tests
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Statistical Tests")
ws4.cell(row=1, column=1,
         value="Statistical Tests  |  dbmdz/german-gpt2  |  Collection-level unit  |  2026-04-28"
         ).font = Font(bold=True, size=12)
ws4.cell(row=1, column=1).fill = C_BANNER
ws4.cell(row=1, column=1).font = Font(color="E6B85A", bold=True, size=11)

# Kruskal-Wallis across all collections (min n=3)
coll_groups = {c: grp for c, grp in df.groupby("collection") if len(grp) >= 3}

row = 3
section_header(ws4, row, 1, f"Kruskal-Wallis H  across {len(coll_groups)} collections (n ≥ 3) — one test per metric")
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1

for ci, h in enumerate(["Metric", "H statistic", "p-value", "Significant (p<0.05)", "N collections", "Total N"], 1):
    set_header(ws4.cell(row=row, column=ci), C_ENTR if ci > 1 else C_META, h)
row += 1

for m in METRICS:
    groups = [g[m].dropna().values for g in coll_groups.values() if g[m].notna().sum() >= 3]
    if len(groups) >= 2:
        H, p = sp_stats.kruskal(*groups)
        sig = "YES" if p < 0.05 else "no"
        n_colls = len(groups)
        n_total = sum(len(g) for g in groups)
        for ci, val in enumerate([METRIC_LABELS[m], round(H,3), round(p,4), sig, n_colls, n_total], 1):
            c = ws4.cell(row=row, column=ci, value=val)
            c.font = DATA_FONT
            if ci == 4:
                c.fill = fill("C8E6C9") if sig == "YES" else fill("FFCDD2")
        row += 1

row += 2

# Pairwise Mann-Whitney for S² Mean (top 15 collections by N)
top_colls = df.groupby("collection").size().nlargest(15).index.tolist()

section_header(ws4, row, 1,
    f"Pairwise Mann-Whitney U  — S² Mean  —  top {len(top_colls)} collections by poem count")
ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(top_colls)+2)
row += 1

# Column headers
ws4.cell(row=row, column=1, value="Collection").font = HEADER_FONT
ws4.cell(row=row, column=1).fill = C_META
ws4.cell(row=row, column=2, value="N").font = HEADER_FONT
ws4.cell(row=row, column=2).fill = C_META
for ci, c in enumerate(top_colls, 3):
    set_header(ws4.cell(row=row, column=ci), C_ENTR, c[:20])
row += 1

tc_data = {c: df[df["collection"]==c]["s2_mean"].dropna().values for c in top_colls}

for ri_off, c1 in enumerate(top_colls):
    ws4.cell(row=row+ri_off, column=1, value=c1[:30]).font = DATA_FONT
    ws4.cell(row=row+ri_off, column=2, value=len(tc_data[c1])).font = DATA_FONT
    for ci, c2 in enumerate(top_colls, 3):
        if c1 == c2:
            cell = ws4.cell(row=row+ri_off, column=ci, value="—")
            cell.fill = fill("E0E0E0")
        elif len(tc_data[c1]) < 3 or len(tc_data[c2]) < 3:
            ws4.cell(row=row+ri_off, column=ci, value="")
        else:
            _, p = sp_stats.mannwhitneyu(tc_data[c1], tc_data[c2], alternative="two-sided")
            n1, n2 = len(tc_data[c1]), len(tc_data[c2])
            U_stat = sp_stats.mannwhitneyu(tc_data[c1], tc_data[c2], alternative="two-sided").statistic
            r_rb = 1 - (2*U_stat)/(n1*n2)   # rank-biserial correlation (effect size)
            cell = ws4.cell(row=row+ri_off, column=ci,
                            value=f"p={p:.3f}\nr={r_rb:.2f}")
            cell.font = Font(size=8)
            cell.alignment = Alignment(wrap_text=True)
            if p < 0.001:
                cell.fill = fill("1A5276")
                cell.font = Font(size=8, color="FFFFFF")
            elif p < 0.01:
                cell.fill = fill("2980B9")
                cell.font = Font(size=8, color="FFFFFF")
            elif p < 0.05:
                cell.fill = fill("AED6F1")
            else:
                cell.fill = fill("FDFEFE")

autowidth(ws4, max_w=20)
ws4.column_dimensions["A"].width = 32


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Correlation
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Correlation")
ws5.cell(row=1, column=1, value="Spearman Correlation Matrix — all 2582 poems").font = Font(bold=True, size=12)
ws5.cell(row=1, column=1).fill = C_BANNER
ws5.cell(row=1, column=1).font = Font(color="E6B85A", bold=True, size=11)

corr_cols = METRICS + ["n_lines", "closure", "temporal_year"]
corr_labels = [METRIC_LABELS.get(c, c) for c in corr_cols]

df_corr = df[corr_cols].copy()
df_corr["closure"]       = pd.to_numeric(df_corr["closure"],       errors="coerce")
df_corr["temporal_year"] = pd.to_numeric(df_corr["temporal_year"], errors="coerce")
corr_mat = df_corr.corr(method="spearman").round(3)

# Row/col headers
for ci, lbl in enumerate(corr_labels, 2):
    set_header(ws5.cell(row=2, column=ci), C_ENTR, lbl)
for ri, lbl in enumerate(corr_labels, 3):
    set_header(ws5.cell(row=ri, column=1), C_ENTR, lbl)

# Diverging color scale: dark blue=+1, white=0, dark red=-1
def corr_fill(r):
    if pd.isna(r): return fill("E0E0E0")
    r = max(-1, min(1, r))
    if r > 0:
        g = int(255 - r*150)
        b = int(255 - r*150)
        return fill(f"{255:02X}{g:02X}{b:02X}")
    else:
        r_abs = abs(r)
        g = int(255 - r_abs*150)
        r_val = 255
        return fill(f"{r_val:02X}{g:02X}{g:02X}")

for ri, row_m in enumerate(corr_cols, 3):
    for ci, col_m in enumerate(corr_cols, 2):
        val = corr_mat.loc[row_m, col_m]
        cell = ws5.cell(row=ri, column=ci, value=float(val) if not pd.isna(val) else "")
        cell.font = Font(size=9, bold=(row_m == col_m))
        cell.fill = corr_fill(val)
        cell.alignment = Alignment(horizontal="center")

autowidth(ws5, max_w=16)
ws5.column_dimensions["A"].width = 18


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Temporal Confound
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Temporal Confound")
ws6.cell(row=1, column=1,
         value="Temporal Confound Analysis — S² Mean and Mean Surprisal by Decade"
         ).fill = C_BANNER
ws6.cell(row=1, column=1).font = Font(color="E6B85A", bold=True, size=11)

ws6.cell(row=3, column=1,
         value=("NOTE: temporal_year = Erstdruck (first publication) when available (853/2582 poems), "
                "else comp_year midpoint when date_uncertain=False (range ≤40 yrs). "
                "Lifespan-proxy comp dates (range >40 yrs, e.g. Goethe 1749-1832) are excluded. "
                "This yields ~1,400-1,600 poems with reliable dates. "
                "Erstdruck was preferred as it reflects actual first publication; comp_year alone had "
                "58% unreliable values (lifespan proxies). Both columns are in the Poems sheet.")
         ).font = Font(italic=True, size=9, color="7F8C8D")
ws6.cell(row=3, column=1).alignment = Alignment(wrap_text=True)
ws6.row_dimensions[3].height = 55

dec_df = df[df["temporal_decade"].notna()].copy()
dec_df["temporal_decade"] = dec_df["temporal_decade"].astype(int)
dec_agg = dec_df.groupby("temporal_decade").agg(
    n=("poem_id", "count"),
    **{f"{m}_mean": (m, "mean") for m in METRICS},
    **{f"{m}_sd":   (m, "std")  for m in METRICS},
).round(4).reset_index()

dec_cols = (
    [("Decade", "temporal_decade", C_DATE), ("N", "n", C_STRUCT)]
    + [(f"{METRIC_LABELS[m]} Mean", f"{m}_mean", C_ENTR) for m in METRICS]
    + [(f"{METRIC_LABELS[m]} SD",   f"{m}_sd",   C_ENTR) for m in METRICS]
)

row = 5
for ci, (label, _, cf) in enumerate(dec_cols, 1):
    set_header(ws6.cell(row=row, column=ci), cf, label)
row += 1

for _, rdata in dec_agg.iterrows():
    row_fill = fill("F0F4F8")
    for ci, (_, col_key, _) in enumerate(dec_cols, 1):
        val = rdata.get(col_key, "")
        if pd.isna(val): val = ""
        cell = ws6.cell(row=row, column=ci, value=val)
        cell.font = DATA_FONT
        cell.fill = row_fill
    row += 1

# Spearman r between temporal_year and metrics
row += 2
section_header(ws6, row, 1, "Spearman r between temporal_year and metrics (poems with valid temporal_year)")
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

for ci, h in enumerate(["Metric", "Spearman r", "p-value", "N"], 1):
    set_header(ws6.cell(row=row, column=ci), C_ENTR if ci > 1 else C_META, h)
row += 1

yr_df = df[df["temporal_year"].notna()].copy()
for m in METRICS:
    sub = yr_df[[m, "temporal_year"]].dropna()
    if len(sub) >= 10:
        r, p = sp_stats.spearmanr(sub[m], sub["temporal_year"])
        for ci, val in enumerate([METRIC_LABELS[m], round(r,4), round(p,4), len(sub)], 1):
            c = ws6.cell(row=row, column=ci, value=val)
            c.font = DATA_FONT
            if ci == 3 and isinstance(val, float):
                c.fill = fill("C8E6C9") if val < 0.05 else fill("FDFEFE")
        row += 1

autowidth(ws6)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Top Outliers
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Top Outliers")
ws7.cell(row=1, column=1,
         value="Top Outliers — per metric  |  Top 15 highest + 15 lowest for each"
         ).fill = C_BANNER
ws7.cell(row=1, column=1).font = Font(color="E6B85A", bold=True, size=11)

row = 3
for metric in METRICS:
    metric_lbl = METRIC_LABELS[metric]
    sub = df[df[metric].notna()].copy()

    section_header(ws7, row, 1, f"Top Outliers — {metric_lbl}")
    ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    for ci, h in enumerate(["Rank", "Poet", "Collection", "Title", metric_lbl, "Mean Surprisal"], 1):
        cf = C_ENTR if ci > 4 else C_META
        set_header(ws7.cell(row=row, column=ci), cf, h)
    row += 1

    section_header(ws7, row, 1, f"HIGHEST {metric_lbl}", bgcolor="154360", fgcolor="ECF0F1")
    row += 1
    for rank, (_, rdata) in enumerate(sub.nlargest(15, metric).iterrows(), 1):
        for ci, val in enumerate([
            rank, rdata["poet"], rdata["collection"][:30],
            rdata["poem_title"][:35], round(rdata[metric], 4),
            round(rdata["mean_surprisal"], 4) if pd.notna(rdata.get("mean_surprisal")) else ""
        ], 1):
            c = ws7.cell(row=row, column=ci, value=val)
            c.font = DATA_FONT
            c.fill = fill("EAF2FF")
        row += 1

    section_header(ws7, row, 1, f"LOWEST {metric_lbl}", bgcolor="6E2F0E", fgcolor="ECF0F1")
    row += 1
    for rank, (_, rdata) in enumerate(sub.nsmallest(15, metric).iterrows(), 1):
        for ci, val in enumerate([
            rank, rdata["poet"], rdata["collection"][:30],
            rdata["poem_title"][:35], round(rdata[metric], 4),
            round(rdata["mean_surprisal"], 4) if pd.notna(rdata.get("mean_surprisal")) else ""
        ], 1):
            c = ws7.cell(row=row, column=ci, value=val)
            c.font = DATA_FONT
            c.fill = fill("FFF0E8")
        row += 1

    row += 2

autowidth(ws7)
ws7.column_dimensions["C"].width = 35
ws7.column_dimensions["D"].width = 40


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8: Notes
# ═══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Notes")
ws8.cell(row=1, column=1, value="Notes and Caveats — v5").fill = C_BANNER
ws8.cell(row=1, column=1).font = Font(color="E6B85A", bold=True, size=12)

NOTES = [
    ("1. Model", "GPT-2 surprisal computed with dbmdz/german-gpt2. One forward pass per poem "
     "(truncated at 900 tokens). Word surprisal = sum of subword token surprisals. "
     "Mean entropy = next-word entropy at first subword position of each word."),
    ("2. S² Mean", "Excess surprisal = word surprisal minus entropy. Measures deviation of actual "
     "word choice beyond contextual predictability (Fraser 2025)."),
    ("3. UID σ", "Std dev of per-word surprisal. Probe of Uniform Information Density (Jaeger 2010). "
     "Low = smooth distribution; high = spiky."),
    ("4. Tension", "Mean surprisal of poem's second half minus first half. Positive = poem "
     "builds toward linguistic difficulty."),
    ("5. AC1", "Lag-1 autocorrelation of surprisal sequence. High = metrically regular."),
    ("6. Annotations", "All 10 literary annotation fields produced by claude-haiku-4-5-20251001 "
     "using a PEPP-grounded coding manual (Ramazani et al., Princeton Encyclopedia of Poetry "
     "and Poetics, 4th ed., 2012). Person field is systematically missing from model output."),
    ("7. George Normalization", "Stefan George (1868-1933) wrote in deliberate all-lowercase. "
     "The 'George: Normalized' column applies standard German capitalization rules (nouns, "
     "proper names, sentence beginnings) using spaCy de_core_news_lg POS tagging. "
     "GPT-2 surprisal scores were computed on original (lowercase) text."),
    ("8. Corpus", "Source: TextGrid Repository XML (TEI P5). New v5 poets: Albrech von Haller, "
     "Barthold Heinrich Brockes, Droste-Hülshoff, Felix Dörmann, Hölderlin, Mörike, "
     "Morgenstern, Schiller, Stefan George. Nine v4 poets reuse prior metrics (XML files "
     "byte-for-byte identical)."),
    ("9. Exclusions", f"Poems excluded: fewer than 4 lines (fragments), more than 400 lines "
     f"(cycles), theatrical content (TEI <sp>/<speaker> tags, except West-östlicher Divan), "
     f"generic section titles. Total: 2,582 poems, 18 poets."),
    ("10. Collection Titles", "Collection names taken from 5th path segment of TEI n= attribute: "
     "/Literatur/X/Author/Gedichte/CollectionTitle/[poem]. This yields actual published "
     "collection titles rather than generic 'Gedichte'."),
]

row = 3
for label, text in NOTES:
    c1 = ws8.cell(row=row, column=1, value=label)
    c1.font = Font(bold=True, size=10)
    c1.fill = fill("EBF5FB")
    c2 = ws8.cell(row=row, column=2, value=text)
    c2.font = Font(size=9)
    c2.alignment = Alignment(wrap_text=True)
    ws8.row_dimensions[row].height = 45
    row += 1

ws8.column_dimensions["A"].width = 24
ws8.column_dimensions["B"].width = 100


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9: Normalization (new vs v4)
# ═══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Normalization")
ws9.cell(row=1, column=1,
         value="Normalization — New Poets vs v4 Corpus  |  George Original vs Normalized"
         ).fill = C_BANNER
ws9.cell(row=1, column=1).font = Font(color="E6B85A", bold=True, size=11)

# v4 poets (those present in v4 surprisal CSV)
V4_POETS = {
    "Albrech von Haller", "Barthold Heinrich Brockes", "Droste-Hülshoff", "Felix Dörmann",
    "Holderlin", "Klopstock", "Morgenstern", "Mörike", "Rilke", "Schiller",
    "goethe", "heym", "hvh", "stadler", "storm", "trakl", "wedekind"
}
NEW_POETS = {"george"}

df["corpus_version"] = df["poet"].apply(lambda p: "v4" if p in V4_POETS else "new (v5)")

row = 3
section_header(ws9, row, 1, "Metric summary: v4 poets (n=17) vs new poet George")
ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
row += 1

norm_cols = ["Corpus", "N", "Metric", "Mean", "Median", "SD", "Min", "Max", "Skew", "Kurt"]
for ci, h in enumerate(norm_cols, 1):
    set_header(ws9.cell(row=row, column=ci), C_ENTR if ci > 2 else C_META, h)
row += 1

for corpus_label, corpus_df in [("v4 poets", df[df["poet"].isin(V4_POETS)]),
                                  ("george (v5)",   df[df["poet"] == "george"])]:
    for m in METRICS:
        sub = corpus_df[m].dropna()
        if len(sub) < 3:
            continue
        vals = [
            corpus_label, len(sub), METRIC_LABELS[m],
            round(sub.mean(), 4), round(sub.median(), 4),
            round(sub.std(), 4), round(sub.min(), 4), round(sub.max(), 4),
            round(float(sp_stats.skew(sub)), 4), round(float(sp_stats.kurtosis(sub)), 4),
        ]
        row_fill = fill("EAF2FF") if "v4" in corpus_label else fill("FFF3E0")
        for ci, val in enumerate(vals, 1):
            c = ws9.cell(row=row, column=ci, value=val)
            c.font = DATA_FONT
            c.fill = row_fill
        row += 1

# Mann-Whitney: george vs v4 pool
row += 2
section_header(ws9, row, 1, "Mann-Whitney U: George vs v4 poets (per metric)")
ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

for ci, h in enumerate(["Metric", "U statistic", "p-value", "Rank-biserial r", "Significant"], 1):
    set_header(ws9.cell(row=row, column=ci), C_ENTR if ci > 1 else C_META, h)
row += 1

geo_df = df[df["poet"] == "george"]
v4_df  = df[df["poet"].isin(V4_POETS)]
for m in METRICS:
    g = geo_df[m].dropna()
    v = v4_df[m].dropna()
    if len(g) >= 3 and len(v) >= 3:
        U, p = sp_stats.mannwhitneyu(g, v, alternative="two-sided")
        r_rb = 1 - (2*U) / (len(g)*len(v))
        sig = "YES" if p < 0.05 else "no"
        for ci, val in enumerate([METRIC_LABELS[m], round(float(U),1), round(p,4), round(r_rb,4), sig], 1):
            c = ws9.cell(row=row, column=ci, value=val)
            c.font = DATA_FONT
            if ci == 5:
                c.fill = fill("C8E6C9") if sig == "YES" else fill("FDFEFE")
        row += 1

# George sample: original vs normalized text
row += 2
section_header(ws9, row, 1, "George: Original vs Normalized Text (first 20 poems)")
ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

for ci, h in enumerate(["Poem Title", "Collection", "Original Text (first line)", "Normalized (first line)"], 1):
    cf = C_DERIV if ci > 2 else C_META
    set_header(ws9.cell(row=row, column=ci), cf, h)
row += 1

george_poems = df[df["poet"] == "george"].head(20)
for _, rdata in george_poems.iterrows():
    orig_line = str(rdata.get("text", "")).split("\n")[0][:100]
    norm_line = str(rdata.get("george_normalized_text", "")).split("\n")[0][:100]
    for ci, val in enumerate([rdata["poem_title"], rdata["collection"][:40], orig_line, norm_line], 1):
        c = ws9.cell(row=row, column=ci, value=val)
        c.font = DATA_FONT
        c.fill = fill("FFF3E0")
    row += 1

autowidth(ws9, max_w=60)
ws9.column_dimensions["A"].width = 30
ws9.column_dimensions["C"].width = 55
ws9.column_dimensions["D"].width = 55


# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
out_path = OUT_DIR / "german_poetry_v5.xlsx"
wb.save(out_path)
print(f"Saved → {out_path}")
print(f"Sheets: {wb.sheetnames}")
