"""
03_make_workbook.py  —  Build Excel workbook with 4 sheets.

Input:   output/poems_metrics_v2.csv, output/exclusions_v2.csv
Output:  output/german_poetry_v2.xlsx

Sheets:
  1. Poems           — poem-level metrics + metadata; red gradient on high surprisal/entropy/S2
  2. Collections     — mean ± SD per collection; red gradient
  3. Exclusions      — excluded poems + ambiguity flags
  4. Temporal        — first pub year as confounder analysis (Spearman + scatter data)
"""

from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats

OUT_DIR = Path.home() / "german-poetry-v5" / "output"

df      = pd.read_csv(OUT_DIR / "poems_metrics_v2.csv")
excl_df = pd.read_csv(OUT_DIR / "exclusions_v2.csv")

wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ── Palette ──────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")
HDR_FONT  = Font(color="E6B85A", bold=True, size=10)
ALT_FILL  = PatternFill("solid", fgColor="F5F5F0")
PLAIN_FONT = Font(size=10)
BOLD_FONT  = Font(bold=True, size=10)
THIN      = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def write_header(ws, cols):
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN

def write_df(ws, frame, alt_rows=True):
    write_header(ws, list(frame.columns))
    for ri, row in enumerate(frame.itertuples(index=False), 2):
        fill = ALT_FILL if (alt_rows and ri % 2 == 0) else PatternFill()
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=None if pd.isna(val) else val)
            cell.font  = PLAIN_FONT
            cell.border = THIN
            cell.fill  = fill

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def add_color_scale(ws, col_letter, min_row, max_row,
                    lo="63BE7B", mid="FFEB84", hi="F8696B"):
    ws.conditional_formatting.add(
        f"{col_letter}{min_row}:{col_letter}{max_row}",
        ColorScaleRule(
            start_type="min", start_color=lo,
            mid_type="percentile", mid_value=50, mid_color=mid,
            end_type="max", end_color=hi,
        )
    )


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Poems
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Poems")

POEM_COLS = [
    "poem_id", "poet", "poem_title", "collection_title",
    "collection_pub_year", "collection_pub_year_uncertain", "collection_pub_year_note",
    "source_file", "total_verse_lines", "stanza_pattern", "lines_used_for_metrics",
    "mean_token_surprisal", "mean_token_entropy", "s2", "n_tokens_used",
    "model", "compute_error",
]
poems_out = df[[c for c in POEM_COLS if c in df.columns]].sort_values(
    ["poet", "collection_pub_year", "poem_title"], na_position="last"
)

write_df(ws1, poems_out)
n_data = len(poems_out) + 1  # +1 for header

# Color-scale on surprisal, entropy, s2
for metric_col in ["mean_token_surprisal", "mean_token_entropy", "s2"]:
    if metric_col in poems_out.columns:
        idx = list(poems_out.columns).index(metric_col) + 1
        cl  = get_column_letter(idx)
        add_color_scale(ws1, cl, 2, n_data)

set_col_widths(ws1, {
    "A": 8, "B": 22, "C": 38, "D": 30,
    "E": 8, "F": 10, "G": 40, "H": 35,
    "I": 8, "J": 14, "K": 8,
    "L": 10, "M": 10, "N": 10, "O": 8,
    "P": 28, "Q": 30,
})
ws1.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Collections
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Collections")

df_ok = df[df["mean_token_surprisal"].notna()].copy()

coll_agg = (
    df_ok.groupby(["poet", "collection_title"])
    .agg(
        collection_pub_year    =("collection_pub_year",           "first"),
        year_uncertain         =("collection_pub_year_uncertain",  "first"),
        year_note              =("collection_pub_year_note",       "first"),
        n_poems                =("poem_id",                        "count"),
        mean_surprisal         =("mean_token_surprisal",           "mean"),
        sd_surprisal           =("mean_token_surprisal",           "std"),
        mean_entropy           =("mean_token_entropy",             "mean"),
        sd_entropy             =("mean_token_entropy",             "std"),
        mean_s2                =("s2",                             "mean"),
        sd_s2                  =("s2",                             "std"),
        mean_lines             =("total_verse_lines",              "mean"),
    )
    .round(4)
    .reset_index()
    .sort_values(["collection_pub_year", "poet"], na_position="last")
)

write_df(ws2, coll_agg)
n2 = len(coll_agg) + 1
for mc, ci in [("mean_surprisal", 8), ("mean_entropy", 10), ("mean_s2", 12)]:
    cl = get_column_letter(ci)
    add_color_scale(ws2, cl, 2, n2)

set_col_widths(ws2, {
    "A": 22, "B": 32, "C": 8, "D": 8, "E": 40,
    "F": 8, "G": 10, "H": 10, "I": 10, "J": 10,
    "K": 10, "L": 10, "M": 10, "N": 10,
})
ws2.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Exclusions
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Exclusions")

# Also add uncertainty-flagged poems as an annotation section
uncertain = df[df["collection_pub_year_uncertain"] == True][
    ["poem_id","poet","collection_title","collection_pub_year","collection_pub_year_note"]
].copy()
uncertain.insert(0, "flag_type", "pub_year_uncertain")
uncertain.rename(columns={
    "collection_title":  "detail",
    "collection_pub_year": "year",
    "collection_pub_year_note": "note",
}, inplace=True)

# Combine exclusions with uncertainty flags
excl_out = excl_df[["source_file","poet","poem_title","reason"]].copy()
write_df(ws3, excl_out)

# Blank row then uncertainty section
blank_row = len(excl_out) + 3
ws3.cell(row=blank_row, column=1, value="⚑ PUB-YEAR UNCERTAINTY FLAGS").font = BOLD_FONT
write_header(ws3, list(uncertain.columns))
# Shift header row
target_row = blank_row + 1
write_header_at = lambda frame, row: [
    ws3.cell(row=row, column=ci+1, value=col).fill == HDR_FILL
    for ci, col in enumerate(frame.columns)
]
for ci, col in enumerate(uncertain.columns, 1):
    cell = ws3.cell(row=target_row, column=ci, value=col)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.border = THIN
for ri, row in enumerate(uncertain.itertuples(index=False), target_row + 1):
    for ci, val in enumerate(row, 1):
        cell = ws3.cell(row=ri, column=ci, value=None if pd.isna(val) else val)
        cell.font  = PLAIN_FONT
        cell.border = THIN

set_col_widths(ws3, {"A": 35, "B": 22, "C": 40, "D": 45})
ws3.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Temporal analysis
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Temporal Analysis")

df_dated = df_ok[df_ok["collection_pub_year"].notna()].copy()
df_dated["collection_pub_year"] = df_dated["collection_pub_year"].astype(int)
df_certain = df_dated[df_dated["collection_pub_year_uncertain"] == False]

metrics = ["mean_token_surprisal", "mean_token_entropy", "s2"]

results = []
for subset_name, subset in [("All dated poems", df_dated),
                             ("Certain years only", df_certain)]:
    for m in metrics:
        sub = subset[["collection_pub_year", m]].dropna()
        if len(sub) < 10:
            continue
        r, p = stats.spearmanr(sub["collection_pub_year"], sub[m])
        results.append({
            "subset":  subset_name,
            "metric":  m,
            "n":       len(sub),
            "spearman_r":  round(r, 4),
            "p_value":     round(p, 6),
            "sig":         "***" if p < .001 else "**" if p < .01 else "*" if p < .05 else "",
            "interpretation": (
                "Positive: metric increases over time" if r > 0 else
                "Negative: metric decreases over time"
            ),
        })

res_df = pd.DataFrame(results)

ws4.cell(row=1, column=1,
         value="Temporal Analysis: first publication year of original collection vs. metrics"
         ).font = BOLD_FONT
ws4.cell(row=2, column=1,
         value="NOTE: Only first pub year of original collection used (not Sammelband/edition year)."
         ).font = Font(italic=True, size=10)
ws4.cell(row=3, column=1,
         value="'Certain years only' subset excludes all poems where collection_pub_year_uncertain=True."
         ).font = Font(italic=True, size=10)

write_header(ws3, [])  # no-op — just writing res_df below
for ci, col in enumerate(res_df.columns, 1):
    cell = ws4.cell(row=5, column=ci, value=col)
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.border = THIN

for ri, row in enumerate(res_df.itertuples(index=False), 6):
    for ci, val in enumerate(row, 1):
        cell = ws4.cell(row=ri, column=ci, value=None if pd.isna(val) else val)
        cell.font  = PLAIN_FONT
        cell.border = THIN

# Scatter data for reference
ws4.cell(row=len(results) + 9, column=1,
         value="Scatter data (collection means by pub year):").font = BOLD_FONT

coll_means = (
    df_dated.groupby(["poet","collection_title","collection_pub_year"])[metrics]
    .mean().round(4).reset_index()
    .sort_values("collection_pub_year")
)
scatter_row = len(results) + 10
for ci, col in enumerate(coll_means.columns, 1):
    cell = ws4.cell(row=scatter_row, column=ci, value=col)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = THIN
for ri, row in enumerate(coll_means.itertuples(index=False), scatter_row + 1):
    for ci, val in enumerate(row, 1):
        cell = ws4.cell(row=ri, column=ci, value=None if pd.isna(val) else val)
        cell.font = PLAIN_FONT; cell.border = THIN

set_col_widths(ws4, {"A": 22, "B": 25, "C": 8, "D": 12, "E": 12, "F": 8, "G": 45})


# ── Save ──────────────────────────────────────────────────────────────────────
out_path = OUT_DIR / "german_poetry_v2.xlsx"
wb.save(out_path)
print(f"Saved → {out_path}")
print(f"  Sheet 1 (Poems):     {len(poems_out)} rows")
print(f"  Sheet 2 (Collections): {len(coll_agg)} collections")
print(f"  Sheet 3 (Exclusions): {len(excl_out)} exclusions + {len(uncertain)} uncertainty flags")
print(f"  Sheet 4 (Temporal):  {len(results)} correlation tests")
